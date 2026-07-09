"""Spreadsheet reader — loads and filters ESA incentive plan data.

Reads the incentive_plan.xlsx workbook produced by the incentive-pipeline.
Provides typed dataclasses and window/stream filters with DST-safe time
comparisons (Europe/Stockholm).
"""

import json
import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from .slugs import stream_token as normalize_stream_token

TZ = ZoneInfo("Europe/Stockholm")


@dataclass
class RunRow:
    pick: int
    scheduled: datetime
    game: str
    category: str
    estimate: str
    platform: str
    players: str
    # Flat fields kept for backwards compatibility — populated from participants[0]
    # (or from the spreadsheet's "Runner"/"Twitch"/etc. columns for legacy sheets).
    runner_display: str
    runner_twitch: str
    runner_discord: str
    runner_twitter: str
    note: Optional[str]
    layout: Optional[str]
    stream: str
    submission_id: Optional[str]
    category_id: Optional[str]
    incentives: str
    commentator: str
    upload_speed: str
    pronouns: str
    show_cam: str
    runner_comments: str
    # Multi-runner list; populated from hidden Participants JSON col when present.
    participants: list = field(default_factory=list)


@dataclass
class IncentiveRow:
    scheduled: datetime
    game: str
    category: str
    stream: str
    # Flat fields kept for backwards compatibility.
    runner_display: str
    runner_twitch: str
    runner_discord: str
    incentive_text: str
    incentive_category: str
    valid_for_game: str
    incentive_estimate: str
    needs_approval: str
    status: str
    submission_id: str
    uuid: str
    participants: list = field(default_factory=list)


@dataclass
class SubmissionRow:
    submission_id: int
    runner: str
    twitch: str
    discord: str
    twitter: str
    game: str
    console: str
    category: str
    estimate: str
    emulated: bool
    incentives: str
    commentator: str
    upload_speed: str
    pronouns: str
    show_cam: str
    comments: str


def _parse_iso_cell(value) -> datetime | None:
    """Parse a datetime cell that may be a string or already a datetime."""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=TZ)
        return value
    if isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=TZ)
            return dt
        except (ValueError, TypeError):
            return None
    return None


def _parse_participants_json(raw) -> list:
    """Safely parse a participants JSON cell. Returns [] on missing/invalid."""
    if not raw:
        return []
    try:
        data = json.loads(str(raw))
        if isinstance(data, list):
            return data
    except (json.JSONDecodeError, TypeError, ValueError):
        pass
    return []


def read_cross_reference(path: str) -> list[RunRow]:
    """Read all runs from the Cross-Reference sheet."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Cross-Reference"]
    rows: list[RunRow] = []
    for i, cell_row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not cell_row or not cell_row[1]:
            continue
        scheduled = _parse_iso_cell(cell_row[0])
        if scheduled is None:
            continue
        # Column 22 (index 21) = Participants JSON (new hidden col)
        participants = _parse_participants_json(cell_row[21] if len(cell_row) > 21 else None)
        # Back-fill single-participant list from flat columns for legacy sheets.
        if not participants:
            runner_display = str(cell_row[6] or "")
            runner_twitch = str(cell_row[7] or "")
            runner_discord = str(cell_row[8] or "")
            runner_twitter = str(cell_row[9] or "")
            if runner_display or runner_twitch:
                participants = [{
                    "display": runner_display,
                    "twitch": runner_twitch,
                    "discord": runner_discord,
                    "twitter": runner_twitter,
                    "pronunciation": "",
                    "submission_id": str(cell_row[13]) if cell_row[13] else None,
                    "match_confidence": "primary",
                }]
        rows.append(RunRow(
            pick=i + 1,
            scheduled=scheduled,
            game=str(cell_row[1] or ""),
            category=str(cell_row[2] or ""),
            estimate=str(cell_row[3] or ""),
            platform=str(cell_row[4] or ""),
            players=str(cell_row[5] or ""),
            runner_display=str(cell_row[6] or ""),
            runner_twitch=str(cell_row[7] or ""),
            runner_discord=str(cell_row[8] or ""),
            runner_twitter=str(cell_row[9] or ""),
            note=str(cell_row[10]) if cell_row[10] else None,
            layout=str(cell_row[11]) if cell_row[11] else None,
            stream=str(cell_row[12] or ""),
            submission_id=str(cell_row[13]) if cell_row[13] else None,
            category_id=str(cell_row[14]) if cell_row[14] else None,
            incentives=str(cell_row[15] or ""),
            commentator=str(cell_row[16] or ""),
            upload_speed=str(cell_row[17] or ""),
            pronouns=str(cell_row[18] or ""),
            show_cam=str(cell_row[19] or ""),
            runner_comments=str(cell_row[20] or ""),
            participants=participants,
        ))
    wb.close()
    return rows


def read_incentives(path: str) -> list[IncentiveRow]:
    """Read all incentive rows from the Incentives Detail sheet."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Incentives Detail"]
    except KeyError:
        wb.close()
        return []
    rows: list[IncentiveRow] = []
    for cell_row in ws.iter_rows(min_row=2, values_only=True):
        if not cell_row or not cell_row[1]:
            continue
        scheduled = _parse_iso_cell(cell_row[0])
        if scheduled is None:
            continue
        # Column 16 (index 15) = UUID; column 17 (index 16) = Participants JSON
        participants = _parse_participants_json(cell_row[16] if len(cell_row) > 16 else None)
        if not participants:
            runner_display = str(cell_row[4] or "")
            runner_twitch = str(cell_row[5] or "")
            runner_discord = str(cell_row[6] or "")
            if runner_display or runner_twitch:
                participants = [{
                    "display": runner_display,
                    "twitch": runner_twitch,
                    "discord": runner_discord,
                    "pronunciation": "",
                    "submission_id": str(cell_row[13] if len(cell_row) > 13 else ""),
                    "match_confidence": "primary",
                }]
        rows.append(IncentiveRow(
            scheduled=scheduled,
            game=str(cell_row[1] or ""),
            category=str(cell_row[2] or ""),
            stream=str(cell_row[3] or ""),
            runner_display=str(cell_row[4] or ""),
            runner_twitch=str(cell_row[5] or ""),
            runner_discord=str(cell_row[6] or ""),
            incentive_text=str(cell_row[7] or ""),
            incentive_category=str(cell_row[8] or ""),
            valid_for_game=str(cell_row[9] or ""),
            incentive_estimate=str(cell_row[10] or ""),
            needs_approval=str(cell_row[11] or ""),
            status=str(cell_row[12] or ""),
            submission_id=str(cell_row[13] if len(cell_row) > 13 else ""),
            uuid=str(cell_row[14] if len(cell_row) > 14 else ""),
            participants=participants,
        ))
    wb.close()
    return rows


def read_submissions(path: str) -> list[SubmissionRow]:
    """Read all submissions from the Submissions sheet."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Submissions"]
    rows: list[SubmissionRow] = []
    for cell_row in ws.iter_rows(min_row=2, values_only=True):
        if not cell_row or not cell_row[1]:
            continue
        rows.append(SubmissionRow(
            submission_id=int(cell_row[0]) if cell_row[0] else 0,
            runner=str(cell_row[1] or ""),
            twitch=str(cell_row[2] or ""),
            discord=str(cell_row[3] or ""),
            twitter=str(cell_row[4] or ""),
            game=str(cell_row[5] or ""),
            console=str(cell_row[6] or ""),
            category=str(cell_row[7] or ""),
            estimate=str(cell_row[8] or ""),
            emulated=str(cell_row[9] or "").lower() == "yes",
            incentives=str(cell_row[10] or ""),
            commentator=str(cell_row[11] or ""),
            upload_speed=str(cell_row[12] or ""),
            pronouns=str(cell_row[13] or ""),
            show_cam=str(cell_row[14] or ""),
            comments=str(cell_row[15] or ""),
        ))
    wb.close()
    return rows


def read_cross_reference_from_db(db_path: str) -> list[RunRow]:
    """Read runs from SQLite DB, returning RunRow dataclasses.

    Enables CLI tools to use ``--source=db`` without changing their
    downstream logic.
    """
    from .db import Run, RunParticipant, make_engine
    from sqlmodel import Session, select as q

    engine = make_engine(db_path)
    rows: list[RunRow] = []
    with Session(engine) as s:
        db_runs = s.exec(q(Run).order_by(Run.pick)).all()
        for r in db_runs:
            scheduled = r.scheduled
            if scheduled.tzinfo is None:
                scheduled = scheduled.replace(tzinfo=TZ)

            participants_raw: list[dict] = []
            db_parts = s.exec(
                q(RunParticipant).where(RunParticipant.run_id == r.id)
            ).all()
            for p in db_parts:
                participants_raw.append({
                    "display": p.display_name,
                    "twitch": p.twitch,
                    "discord": p.discord,
                    "twitter": p.twitter,
                    "pronunciation": p.pronouns or "",
                    "submission_id": p.submission_id,
                    "match_confidence": p.match_confidence,
                })

            p0 = db_parts[0] if db_parts else None
            rows.append(RunRow(
                pick=r.pick,
                scheduled=scheduled,
                game=r.game,
                category=r.category,
                estimate=r.estimate,
                platform=r.platform,
                players=r.players,
                runner_display=p0.display_name if p0 else "",
                runner_twitch=p0.twitch if p0 else "",
                runner_discord=p0.discord if p0 else "",
                runner_twitter=p0.twitter if p0 else "",
                note=r.note,
                layout=r.layout,
                stream=r.stream,
                submission_id=r.submission_id,
                category_id=r.category_id,
                incentives=r.incentives,
                commentator=r.commentator,
                upload_speed=r.upload_speed,
                pronouns=r.pronouns,
                show_cam=r.show_cam,
                runner_comments=r.runner_comments,
                participants=participants_raw,
            ))
    return rows


def read_incentives_from_db(db_path: str) -> list[IncentiveRow]:
    """Read incentives from SQLite DB, returning IncentiveRow dataclasses."""
    from .db import Incentive, make_engine
    from sqlmodel import Session, select as q

    engine = make_engine(db_path)
    rows: list[IncentiveRow] = []
    with Session(engine) as s:
        db_invs = s.exec(q(Incentive).order_by(Incentive.scheduled)).all()
        for inv in db_invs:
            scheduled = inv.scheduled
            if scheduled.tzinfo is None:
                scheduled = scheduled.replace(tzinfo=TZ)

            participants_raw: list[dict] = []
            if inv.participants_json:
                try:
                    parsed = json.loads(inv.participants_json)
                    if isinstance(parsed, list):
                        participants_raw = parsed
                except (json.JSONDecodeError, TypeError, ValueError):
                    pass

            rows.append(IncentiveRow(
                scheduled=scheduled,
                game=inv.game,
                category=inv.category,
                stream=inv.stream,
                runner_display="",
                runner_twitch="",
                runner_discord="",
                incentive_text=inv.incentive_text,
                incentive_category=inv.incentive_category,
                valid_for_game=inv.valid_for_game,
                incentive_estimate=inv.incentive_estimate,
                needs_approval=inv.needs_approval,
                status=inv.status,
                submission_id=inv.submission_id,
                uuid=inv.uuid,
                participants=participants_raw,
            ))
    return rows


def filter_runs_by_window(
    runs: list[RunRow],
    start: datetime,
    end: datetime,
) -> list[RunRow]:
    """Filter runs within [start, end) time window.

    All datetimes are compared as Europe/Stockholm-aware.
    """
    tz = TZ
    start_aware = start if start.tzinfo else start.replace(tzinfo=tz)
    end_aware = end if end.tzinfo else end.replace(tzinfo=tz)
    return [r for r in runs if start_aware <= r.scheduled < end_aware]


def filter_runs_by_stream(runs: list[RunRow], stream_token: str) -> list[RunRow]:
    """Filter runs matching a stream name or token.

    stream_token may be the full stream name (e.g. "2026 - Summer (Stream One)")
    or a short token (e.g. "stream1", "One"). Matches case-insensitive suffix.
    """
    token_lower = stream_token.lower()
    match_all = not token_lower  # empty token = no filter

    def _matches(run: RunRow) -> bool:
        if match_all:
            return True
        stream_lower = run.stream.lower()
        if stream_lower == token_lower:
            return True
        if stream_lower.endswith(token_lower):
            return True
        return False

    return [r for r in runs if _matches(r)]


def stream_token_to_short(stream_name: str) -> str:
    """Normalize a full stream name to a short token. Delegates to
    src.slugs.stream_token for the canonical implementation."""
    return normalize_stream_token(stream_name)


def get_distinct_streams(path: str) -> list[str]:
    """Return ordered list of distinct stream names in the schedule."""
    runs = read_cross_reference(path)
    seen: dict[str, int] = {}
    for r in runs:
        if r.stream not in seen:
            seen[r.stream] = len(seen)
    return sorted(seen, key=lambda s: seen[s])


def check_stale(path: str, max_age_hours: float = 6) -> dict:
    """Check if the spreadsheet is older than max_age_hours.

    Returns dict with: age_hours, is_stale, is_missing.
    """
    if not os.path.exists(path):
        return {"age_hours": None, "is_stale": True, "is_missing": True}

    now = datetime.now(TZ)
    mtime = datetime.fromtimestamp(os.path.getmtime(path), tz=TZ)
    age = (now - mtime).total_seconds() / 3600
    return {
        "age_hours": round(age, 1),
        "is_stale": age > max_age_hours,
        "is_missing": False,
    }


def check_stale_from_db(db_path: str, max_age_hours: float = 6) -> dict:
    """Check if the SQLite DB is older than max_age_hours."""
    if not os.path.exists(db_path):
        return {"age_hours": None, "is_stale": True, "is_missing": True}

    now = datetime.now(TZ)
    mtime = datetime.fromtimestamp(os.path.getmtime(db_path), tz=TZ)
    age = (now - mtime).total_seconds() / 3600
    return {
        "age_hours": round(age, 1),
        "is_stale": age > max_age_hours,
        "is_missing": False,
    }


def find_runner_sibling_runs(
    runs: list[RunRow],
    twitch: str,
    display_name: str,
    exclude_submission_id: str = "",
    max_count: int = 5,
) -> dict:
    """Find other runs in the schedule by the same runner.

    Checks every participant in every run, so co-runners of race entries
    are now discoverable by their own twitch/display. Matches by Twitch
    handle (case-insensitive) primary, display name fallback when Twitch
    is empty. Returns up to max_count chronological upcoming runs plus a
    total count.
    """
    twitch_lower = twitch.strip().lower()
    name_lower = display_name.strip().lower()

    def _run_involves_runner(r: RunRow) -> bool:
        for p in r.participants:
            p_twitch = (p.get("twitch") or "").strip().lower()
            p_display = (p.get("display") or "").strip().lower()
            if twitch_lower and p_twitch == twitch_lower:
                return True
            if not twitch_lower and name_lower and p_display == name_lower:
                return True
        # Fallback to flat fields for legacy rows with no participants list
        if not r.participants:
            r_twitch = r.runner_twitch.strip().lower()
            r_name = r.runner_display.strip().lower()
            if twitch_lower and r_twitch == twitch_lower:
                return True
            if not twitch_lower and r_name and r_name == name_lower:
                return True
        return False

    siblings: list[RunRow] = []
    for r in runs:
        if r.submission_id == exclude_submission_id:
            continue
        if _run_involves_runner(r):
            siblings.append(r)

    now = datetime.now(TZ)
    upcoming = sorted(
        [s for s in siblings if s.scheduled >= now],
        key=lambda x: x.scheduled,
    )
    completed = sorted(
        [s for s in siblings if s.scheduled < now],
        key=lambda x: x.scheduled,
        reverse=True,
    )

    capped = upcoming[:max_count]
    result = {
        "total": len(upcoming),
        "completed_count": len(completed),
        "runs": [
            {
                "pick": r.pick,
                "scheduled": r.scheduled.isoformat(),
                "game": r.game,
                "category": r.category,
                "estimate": r.estimate,
                "stream": stream_token_to_short(r.stream),
                "is_next": i == 0,
            }
            for i, r in enumerate(capped)
        ],
    }
    return result
