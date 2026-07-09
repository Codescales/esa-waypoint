"""Evaluate incentive lookup quality against known game incentives.

Usage:
    python3 -m src.evaluate              # Run evaluation from frozen fixture
    python3 -m src.evaluate --refresh    # Regenerate fixture from spreadsheet first
    python3 -m src.evaluate --emit-manifest  # Emit web-search to-do for unreachable games
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .find_incentives import (
    GDQ_EVENTS, fetch_event_bids, search_bids_for_game, _categorize_bid,
    find_incentives,
)
from .src_api import (
    search_src_game, fetch_src_categories, fetch_category_wr,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
FIXTURE_PATH = PROJECT_ROOT / "tests" / "eval_dataset.json"
REPORT_PATH = PROJECT_ROOT / "tests" / "eval_report.json"
MANIFEST_PATH = PROJECT_ROOT / "tests" / "websearch_todo.json"
DEFAULT_CACHE_PATH = PROJECT_ROOT / "tests" / "websearch_cache.json"

# ── helpers ──────────────────────────────────────────────────────────

def _parse_estimate(est: str) -> int:
    if not est:
        return 0
    parts = [int(p) for p in est.split(":")]
    if len(parts) == 3:
        return parts[0] * 60 + parts[1] + parts[2] // 60
    elif len(parts) == 2:
        return parts[0] + parts[1] // 60
    return 0


def _normalize(text: str) -> str:
    text = text.lower()
    text = text.replace("é", "e").replace("è", "e").replace("ê", "e").replace("ë", "e")
    text = text.replace("à", "a").replace("â", "a").replace("ä", "a")
    text = text.replace("ù", "u").replace("û", "u").replace("ü", "u")
    text = text.replace("ô", "o").replace("ö", "o")
    text = text.replace("î", "i").replace("ï", "i")
    text = text.replace("ç", "c")
    text = re.sub(r"&#?\w+;", "", text)
    text = re.sub(r"[&\+\-\u2014\u2013\'\"]", " ", text)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


STOP_WORDS = {"the", "a", "an", "and", "or", "but", "in", "on", "at", "to", "for",
              "of", "with", "by", "from", "is", "are", "was", "were", "be", "been",
              "being", "have", "has", "had", "do", "does", "did", "will", "would",
              "could", "should", "may", "might", "can", "shall", "this", "that",
              "these", "those", "it", "its", "we", "our", "you", "your", "they",
              "them", "their", "he", "she", "him", "her", "his", "not", "no", "nor",
              "so", "as", "if", "then", "than", "too", "very", "just", "about",
              "also", "into", "over", "after", "before", "between", "through",
              "during", "because", "when", "where", "how", "what", "which", "who",
              "there", "up", "out", "all", "each", "every", "both", "few", "more",
              "most", "some", "any", "one", "two", "three", "first", "second", "last",
              "new", "old", "other", "same", "different", "another", "much", "many",
              "well", "way", "get", "go", "make", "use", "see", "know", "think",
              "come", "take", "give", "find", "tell", "ask", "work", "seem", "feel",
              "try", "leave", "call", "run", "play", "choose", "pick", "option",
              "options", "select", "selection", "add", "added", "adds", "bonus",
              "extra", "additional", "show", "shows", "showcase", "exhibition",
              "display", "demonstrate", "demo", "met", "goal", "target", "goal"}


def _tokenize(text: str) -> set[str]:
    return {t for t in _normalize(text).split() if t not in STOP_WORDS and len(t) > 2}


def _classify_incentive_text(text: str) -> str:
    lower = _normalize(text)
    if any(kw in lower for kw in ["bid war", "choose", "bidwar", "vs ", "selection",
                                    "pick", "choice", "vote", "poll"]):
        return "Poll-Bid War"
    if any(kw in lower for kw in ["if met", "if certain", "donation goal",
                                    "bonus run", "upgrade", "adds ", "target",
                                    "estimate", "if reached", "adds roughly",
                                    "if met"]):
        return "Target"
    return "Reward"


def _semantic_match(actual_text: str, suggestion_text: str) -> bool:
    actual_tokens = _tokenize(actual_text)
    sug_tokens = _tokenize(suggestion_text)
    if not actual_tokens or not sug_tokens:
        return False
    intersection = actual_tokens & sug_tokens
    if not intersection:
        return False
    jaccard = len(intersection) / len(actual_tokens | sug_tokens)
    if jaccard >= 0.2:
        return True
    signal_words = {"nickname", "name", "rename", "language", "costume", "outfit",
                    "skin", "character", "color", "colour", "difficulty", "ending",
                    "route", "weapon", "powerup", "song", "music", "soundtrack",
                    "voice", "texture", "sprite", "model", "handicap", "blindfold",
                    "race", "relay", "coop", "versus", "bid", "cutscene", "glitch",
                    "showcase", "exhibition", "demonstration"}
    shared_signals = intersection & signal_words
    if shared_signals:
        return True
    return False


def _classify_web_snippet(snippet: str) -> str:
    """Infer incentive type from a web snippet."""
    lower = _normalize(snippet)
    if any(kw in lower for kw in ["bid war", "choose", "bidwar", "selection", "choice",
                                    "vote", "poll", "bid", "bidding"]):
        return "Poll-Bid War"
    if any(kw in lower for kw in ["if met", "bonus", "upgrade", "goal", "target",
                                    "showcase", "exhibition", "demonstrate",
                                    "cutscene", "add", "unlock", "bonus game"]):
        return "Target"
    return "Reward"


# ── fixture generation ───────────────────────────────────────────────

def _build_fixture_from_xlsx(path: str) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)

    ws = wb["Incentives Detail"]
    game_inc: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        g = str(row[1]) or ""
        cat = str(row[2]) or ""
        text = str(row[7]) or ""
        valid = str(row[9]) or ""
        status = str(row[12]) or ""
        if not g or not text:
            continue
        if valid == "No":
            continue
        key = f"{g}|{cat}"
        if key not in game_inc:
            game_inc[key] = {"game": g, "category": cat, "actual_incentives": []}
        game_inc[key]["actual_incentives"].append({
            "text": text,
            "type": _classify_incentive_text(text),
            "valid": valid,
            "status": status,
        })

    wb.close()
    wb2 = load_workbook(path, read_only=True, data_only=True)
    ws2 = wb2["Fundraising View"]
    estimates: dict[str, int] = {}
    game_no_inc: list[dict] = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        g = str(row[1]) if row[1] else ""
        cat = str(row[2]) if row[2] else ""
        est = str(row[3]) if row[3] else ""
        inc = str(row[8]) if row[8] else ""
        if not g:
            continue
        key = f"{g}|{cat}"
        if key not in estimates or _parse_estimate(est) > 0:
            estimates[key] = _parse_estimate(est)
        if not inc or inc.lower().strip() in ("nada", "no", "none", ""):
            game_no_inc.append({"game": g, "category": cat, "estimate_minutes": estimates[key]})
    wb2.close()

    confirmation_set = []
    for key, data in game_inc.items():
        est = estimates.get(key, 0)
        confirmation_set.append({
            "game": data["game"],
            "category": data["category"],
            "estimate_minutes": est,
            "actual_incentives": data["actual_incentives"],
        })

    seen_disc = set()
    discovery_set = []
    for d in game_no_inc:
        key = f"{d['game']}|{d['category']}"
        if key not in seen_disc:
            seen_disc.add(key)
            discovery_set.append(d)

    return {
        "generated_from": str(path),
        "generated_at": datetime.now().isoformat() + "Z",
        "confirmation_set": confirmation_set,
        "discovery_set": discovery_set,
    }


# ── web search cache ─────────────────────────────────────────────────

def _get_websearch_cache_path() -> Path:
    env = os.getenv("WEBSEARCH_CACHE")
    if env:
        return Path(env)
    return DEFAULT_CACHE_PATH


def _load_websearch_cache() -> dict:
    """Load web search results cache. Returns {} if absent or unreadable."""
    path = _get_websearch_cache_path()
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text())
    except Exception:
        return {}


# ── evaluation ───────────────────────────────────────────────────────

def _categorize_suggestion(suggestion: dict) -> str:
    return suggestion.get("category", "Reward")


def evaluate_dataset(fixture: dict, gdq_bids: list[dict], web_cache: dict = None) -> dict:
    """Run evaluation for every game in the fixture."""
    if web_cache is None:
        web_cache = {}

    confirmation_set = fixture["confirmation_set"]
    discovery_set = fixture["discovery_set"]

    results: dict = {
        "fixture_generated_at": fixture.get("generated_at", ""),
        "games_confirmed_gdq_src": 0,
        "games_confirmed_with_web": 0,
        "games_missed": 0,
        "games_unreachable": 0,
        "games_not_websearched": 0,
        "incentives_confirmed_type": 0,
        "incentives_confirmed_semantic": 0,
        "incentives_confirmed_semantic_gdq_src": 0,
        "incentives_confirmed_semantic_web": 0,
        "incentives_web_only_rescues": 0,
        "incentives_missed": 0,
        "total_actual_incentives": 0,
        "gdq_total_bids_cached": len(gdq_bids),
        "per_game": [],
        "discovery": [],
    }

    for item in confirmation_set:
        game = item["game"]
        cat = item["category"]
        estimate = item["estimate_minutes"]
        actuals = item["actual_incentives"]

        result = find_incentives(game, cat, estimate, gdq_bids=gdq_bids)
        gdq_suggestions = result.get("incentives", [])
        src_upgrades = result.get("src_upgrades", [])

        web_entries = web_cache.get(game, []) if web_cache else []

        game_result = {
            "game": game,
            "category": cat,
            "estimate_minutes": estimate,
            "gdq_hits": len(gdq_suggestions),
            "src_upgrades": len(src_upgrades),
            "web_entries": len(web_entries),
            "actual_incentives": [],
        }

        any_match_gdq_src = False
        any_match_with_web = False

        for ai in actuals:
            gdq_type = ai["type"]
            text = ai["text"]

            gdq_type_match = False
            gdq_sem_match = False
            src_type_match = False
            src_sem_match = False
            web_type_match = False
            web_sem_match = False

            for sug in gdq_suggestions:
                sug_type = _categorize_suggestion(sug)
                if sug_type == gdq_type:
                    gdq_type_match = True
                    any_reachable = True
                if _semantic_match(text, sug["name"]):
                    gdq_sem_match = True

            for sug in src_upgrades:
                if sug["category"] == "Target" and ai["type"] == "Target":
                    src_type_match = True
                    any_reachable = True
                if _semantic_match(text, sug["name"]):
                    src_sem_match = True

            for entry in web_entries:
                snippet = f"{entry.get('title', '')} {entry.get('snippet', '')}"
                inferred = _classify_web_snippet(snippet)
                if inferred == gdq_type:
                    web_type_match = True
                if _semantic_match(text, snippet):
                    web_sem_match = True

            type_match_gdq_src = gdq_type_match or src_type_match
            sem_match_gdq_src = gdq_sem_match or src_sem_match
            type_match = type_match_gdq_src or web_type_match
            sem_match = sem_match_gdq_src or web_sem_match

            if type_match_gdq_src or sem_match_gdq_src:
                any_match_gdq_src = True
            if type_match or sem_match:
                any_match_with_web = True
            if not type_match_gdq_src and not sem_match_gdq_src and (web_type_match or web_sem_match):
                results["incentives_web_only_rescues"] += 1

            ai_result = {
                "text": text[:80],
                "expected_type": gdq_type,
                "valid": ai.get("valid", ""),
                "type_match": type_match,
                "semantic_match": sem_match,
                "type_match_gdq_src": type_match_gdq_src,
                "semantic_match_gdq_src": sem_match_gdq_src,
                "type_match_web": web_type_match,
                "semantic_match_web": web_sem_match,
            }
            game_result["actual_incentives"].append(ai_result)

            if type_match:
                results["incentives_confirmed_type"] += 1
            if sem_match:
                results["incentives_confirmed_semantic"] += 1
            if sem_match_gdq_src:
                results["incentives_confirmed_semantic_gdq_src"] += 1
            if web_sem_match:
                results["incentives_confirmed_semantic_web"] += 1
            if not type_match and not sem_match:
                results["incentives_missed"] += 1

            results["total_actual_incentives"] += 1

        if any_match_gdq_src:
            results["games_confirmed_gdq_src"] += 1
        if any_match_with_web:
            results["games_confirmed_with_web"] += 1

        has_gdq_src_data = len(gdq_suggestions) > 0 or len(src_upgrades) > 0
        web_cache_active = bool(web_cache)  # cache exists and has entries

        if not any_match_gdq_src and not any_match_with_web:
            if not has_gdq_src_data and not web_entries:
                if web_cache_active:
                    results["games_not_websearched"] += 1
                else:
                    results["games_unreachable"] += 1
            else:
                results["games_missed"] += 1

        results["per_game"].append(game_result)

    # Discovery set
    for item in discovery_set:
        game = item["game"]
        cat = item["category"]
        estimate = item["estimate_minutes"]
        result = find_incentives(game, cat, estimate, gdq_bids=gdq_bids)
        results["discovery"].append({
            "game": game,
            "category": cat,
            "estimate_minutes": estimate,
            "gdq_hits": len(result.get("incentives", [])),
            "gdq_exact": len([i for i in result.get("incentives", [])
                            if game.lower() in i["run"].lower()]),
            "src_upgrades": len(result.get("src_upgrades", [])),
        })

    return results


def _print_report(results: dict):
    total = results["total_actual_incentives"]
    ct = results["incentives_confirmed_type"]
    cs = results["incentives_confirmed_semantic"]
    cs_gs = results["incentives_confirmed_semantic_gdq_src"]
    cs_w = results["incentives_confirmed_semantic_web"]
    cm = results["incentives_missed"]
    rescues = results["incentives_web_only_rescues"]

    print("=" * 60)
    print("  ESA INCENTIVE EVALUATION REPORT")
    print("=" * 60)
    print()

    print(f"Fixture: {results.get('fixture_generated_at', '?')}")
    print(f"GDQ bids cached: {results['gdq_total_bids_cached']}")
    print()

    print("── Semantic match rates (before / after web) ──")
    bp = cs_gs * 100 // total if total else 0
    ap = cs * 100 // total if total else 0
    print(f"  GDQ+SRC only:   {cs_gs}/{total} ({bp}%)")
    print(f"  +Web:           {cs}/{total} ({ap}%)  (now includes web scores)")
    print(f"  Web-only rescues: {rescues}")
    print()

    print("── Type match rates (before / after web) ──")
    ct_gs = results["incentives_confirmed_type"] - results.get("incentives_web_type_only", 0)
    t_bp = ct * 100 // total if total else 0
    print(f"  GDQ+SRC+Web:    {ct}/{total} ({t_bp}%)")
    print()

    print(f"  Missed (neither):     {cm}/{total} ({cm*100//total}%)")
    print()

    gs = results["games_confirmed_gdq_src"]
    gw = results["games_confirmed_with_web"]
    gm = results["games_missed"]
    gu = results["games_unreachable"]
    gnw = results["games_not_websearched"]
    total_gs = gs + gm + gu + gnw
    print(f"── Per-game breakdown ({total_gs} confirmation games) ──")
    print(f"  Confirmed (GDQ+SRC):     {gs}")
    print(f"  Confirmed (with Web):    {gw}")
    print(f"  Missed (data but no hit): {gm}")
    print(f"  Unreachable (no data):     {gu}")
    print(f"  Not yet web-searched:     {gnw}")
    print()

    print("── Per-game details ──")
    for pg in results["per_game"]:
        game = pg["game"]
        gdq = pg["gdq_hits"]
        src = pg["src_upgrades"]
        web = pg["web_entries"]
        incs = pg["actual_incentives"]
        matched_gs = any(i["type_match_gdq_src"] or i["semantic_match_gdq_src"] for i in incs)
        matched_web = any(i["semantic_match"] or i["type_match"] for i in incs)
        tag = "✓" if matched_web else ("~" if matched_gs else ("?" if web == 0 else "✗"))
        print(f"  {tag} {game[:38]:38s} GDQ={gdq:2d} SRC={src:2d} W={web:2d}")
        for ai in incs:
            t = "T" if ai["type_match"] else "·"
            s = "S" if ai["semantic_match"] else "·"
            src_tag = ""
            if ai.get("semantic_match_web") and not ai.get("semantic_match_gdq_src"):
                src_tag = "[web]"
            elif ai.get("semantic_match_web"):
                src_tag = "[both]"
            elif ai.get("semantic_match_gdq_src"):
                src_tag = ""
            print(f"      {t}{s} {src_tag:6s} [{ai['expected_type']}] {ai['text'][:60]}")

    print()
    print(f"── Discovery set ({len(results['discovery'])} games without incentives) ──")
    for d in results["discovery"]:
        game = d["game"]
        gdq = d["gdq_hits"]
        exact = d["gdq_exact"]
        src = d["src_upgrades"]
        print(f"  {game[:40]:40s} GDQ={gdq:2d}(exact={exact}) SRC={src}")
    print()


# ── manifest ─────────────────────────────────────────────────────────

def _build_manifest(fixture: dict, gdq_bids: list[dict], web_cache: dict,
                     force_all: bool = False) -> list[dict]:
    """Build the web-search to-do list for unreachable confirmation games.

    A game is unreachable if GDQ+SRC found nothing. Returns manifest
    entries only for games without a cache entry (or all games if force_all).
    """
    manifest = []
    for item in fixture["confirmation_set"]:
        game = item["game"]
        cat = item["category"]
        estimate = item["estimate_minutes"]

        if not force_all and game in web_cache:
            continue

        result = find_incentives(game, cat, estimate, gdq_bids=gdq_bids)
        gdq_hits = len(result.get("incentives", []))
        src_hits = len(result.get("src_upgrades", []))
        if gdq_hits > 0 or src_hits > 0:
            continue

        query = f'"{game}" speedrun marathon incentive OR bid war OR bonus run'
        manifest.append({
            "game": game,
            "category": cat,
            "estimate_minutes": estimate,
            "query": query,
        })
    return manifest


# ── main ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--refresh", action="store_true", help="Regenerate fixture from spreadsheet")
    parser.add_argument("--emit-manifest", action="store_true",
                        help="Emit web-search to-do manifest for unreachable games")
    parser.add_argument("--refresh-websearch", action="store_true",
                        help="Emit manifest for all unreachable games (ignore existing cache)")
    args = parser.parse_args()

    if args.refresh:
        xlsx_path = PROJECT_ROOT / "output" / "incentive_plan.xlsx"
        if not xlsx_path.exists():
            print(f"ERROR: {xlsx_path} not found. Run the pipeline first.")
            sys.exit(1)
        print("Building fixture from spreadsheet...")
        fixture = _build_fixture_from_xlsx(str(xlsx_path))
        FIXTURE_PATH.write_text(json.dumps(fixture, indent=2))
        print(f"  {len(fixture['confirmation_set'])} confirmation games")
        print(f"  {len(fixture['discovery_set'])} discovery games")
        print(f"  Written to {FIXTURE_PATH}")
    else:
        if not FIXTURE_PATH.exists():
            print(f"ERROR: {FIXTURE_PATH} not found. Run with --refresh first.")
            sys.exit(1)
        fixture = json.loads(FIXTURE_PATH.read_text())
        print(f"Loaded fixture from {FIXTURE_PATH}")
        print(f"  {len(fixture['confirmation_set'])} confirmation games")
        print(f"  {len(fixture['discovery_set'])} discovery games")

    print("Caching GDQ bids...")
    all_gdq_bids = []
    for name, slug in GDQ_EVENTS:
        bids = fetch_event_bids(slug)
        all_gdq_bids.extend(bids)
    print(f"  {len(all_gdq_bids)} parent bids cached")

    web_cache = _load_websearch_cache()
    print(f"  Web search cache: {len(web_cache)} games cached at {_get_websearch_cache_path()}")
    print()

    if args.emit_manifest or args.refresh_websearch:
        force = args.refresh_websearch
        manifest = _build_manifest(fixture, all_gdq_bids, web_cache, force_all=force)
        if not manifest:
            print("All unreachable games have cache entries. Nothing to search.")
        else:
            MANIFEST_PATH.write_text(json.dumps(manifest, indent=2))
            print(f"Emitted manifest with {len(manifest)} games needing web search:")
            for m in manifest:
                print(f"  {m['game']:40s} | {m['query']}")
            print(f"\nWritten to {MANIFEST_PATH}")
            print("Run websearch for each game, then append results to the cache file.")
        return

    results = evaluate_dataset(fixture, all_gdq_bids, web_cache)
    _print_report(results)

    REPORT_PATH.write_text(json.dumps(results, indent=2))
    print(f"Full report written to {REPORT_PATH}")


if __name__ == "__main__":
    main()
