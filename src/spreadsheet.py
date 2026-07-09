"""Spreadsheet generator for ESA incentive planning.

Produces an XLSX workbook with multiple sheets:
  - Schedule: All runs from Horaro
  - Submissions: All Oengus submissions
  - Cross-Reference: Matched schedule items with submission data
  - Incentives Detail: Individual incentives with review workflow
  - Fundraising View: Per-run summary with live formulas + annotations
  - Marathon Info: Summary stats
"""

from dataclasses import dataclass
from datetime import datetime
from typing import Optional
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

import json

from .horaro import ScheduleItem
from .oengus import OengusSubmission, OengusMarathon, find_participant_submissions
from .incentives import IncentiveRow, build_incentive_rows


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


@dataclass
class CrossReferenceRow:
    scheduled: datetime
    game: str
    category: str
    estimate: str
    platform: str
    players: str
    # Flat fields kept for backwards compatibility — populated from participants[0].
    # For multi-runner (race) rows these are joined strings, e.g. "leahkazuno vs mrponytale".
    runner_display: str
    runner_twitch: str
    runner_discord: str
    runner_twitter: str
    note: Optional[str]
    layout: Optional[str]
    stream: str
    submission_id: Optional[str]
    category_id: Optional[str]
    incentives: str
    commentator: str
    upload_speed: str
    pronouns: str
    show_cam: str
    runner_comments: str
    # Multi-runner: list of participant dicts, one per runner.
    # Each dict: {display, twitch, discord, twitter, submission_id, match_confidence}
    participants: list = None  # type: ignore[assignment]
    # All submission IDs linked to this run (primary first).
    runner_submission_ids: list = None  # type: ignore[assignment]

    def __post_init__(self):
        if self.participants is None:
            self.participants = []
        if self.runner_submission_ids is None:
            self.runner_submission_ids = []


def _style_header(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, num_cols: int, alt: bool = False, missing: bool = False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
        if missing:
            cell.fill = MISSING_FILL
        elif alt:
            cell.fill = ALT_FILL


def _auto_width(ws, min_width: int = 8, max_width: int = 40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max(max_len + 2, min_width)


def _find_answer_by_keyword(sub: OengusSubmission, keywords: list[str]) -> str:
    for a in sub.answers:
        label = a.get("label", "").lower()
        for kw in keywords:
            if kw in label:
                return a.get("answer", "")
    return ""


def _union_incentives(subs: list[OengusSubmission]) -> str:
    """Union incentive texts from multiple submissions, deduping by normalised text.

    Normalisation: collapse whitespace, lowercase. Returns the original-case
    text for the first occurrence of each unique normalised form.
    """
    seen: set[str] = set()
    parts: list[str] = []
    for sub in subs:
        raw = _find_answer_by_keyword(sub, ["incentive", "incentivos"])
        if not raw or raw.lower().strip() in ("nada", "no", "none", ""):
            continue
        norm = re.sub(r"\s+", " ", raw).strip().lower()
        if norm not in seen:
            seen.add(norm)
            parts.append(raw.strip())
    return "\n\n".join(parts)


def _build_cross_reference(
    schedule_items: list[ScheduleItem],
    submissions: list[OengusSubmission],
) -> list[CrossReferenceRow]:
    sub_by_id: dict[int, OengusSubmission] = {s.id: s for s in submissions}

    rows = []
    for item in schedule_items:
        # --- Find all participant submissions ---
        matched = find_participant_submissions(item, submissions)
        # matched = [(sub, confidence, key), ...]
        # The primary (Horaro-linked) submission is always first when present.

        participant_subs: list[OengusSubmission] = [m[0] for m in matched]

        # Build participant dicts from matched Oengus users.
        # markdown-only entries (no Oengus match) are appended after.
        participant_dicts: list[dict] = []
        seen_sub_ids: set[int] = set()
        for sub, confidence, key in matched:
            participant_dicts.append({
                "display": sub.user.display_name or sub.user.username or "",
                "twitch": (sub.user.twitch or "").strip().lower(),
                "discord": sub.user.discord or "",
                "twitter": sub.user.twitter or "",
                "pronunciation": "",
                "submission_id": str(sub.id),
                "match_confidence": confidence,
            })
            seen_sub_ids.add(sub.id)

        # Append markdown-only participants (no Oengus submission found).
        # Also backfill twitch from markdown when Oengus user had none.
        for entry in item.participants:
            twitch = (entry.get("twitch") or "").strip().lower()
            display = (entry.get("display") or "").strip().lower()
            # Skip if already accounted for by a matched submission.
            # Cases:
            # 1) Twitch handles match (both non-empty).
            # 2) Display names match (case-insensitive) — covers the common
            #    case where the markdown link text = OengusUser.username but
            #    the Oengus user has no Twitch connection registered.
            matched_p = next(
                (p for p in participant_dicts
                 if (p["twitch"] and twitch and p["twitch"] == twitch)
                 or p["display"].strip().lower() == display),
                None,
            )
            if matched_p is not None:
                # Backfill twitch from markdown if the matched Oengus user had none.
                if twitch and not matched_p["twitch"]:
                    matched_p["twitch"] = twitch
            else:
                participant_dicts.append({
                    "display": entry.get("display") or "",
                    "twitch": twitch,
                    "discord": "",
                    "twitter": "",
                    "pronunciation": "",
                    "submission_id": None,
                    "match_confidence": "markdown-only",
                })

        # If no participants found at all (no submission, no markdown links),
        # leave participants empty — run is unmatched.

        # --- Derive flat backwards-compat fields from participants ---
        if participant_dicts:
            runner_display = " vs ".join(
                p["display"] for p in participant_dicts if p["display"]
            )
            runner_twitch = ", ".join(
                p["twitch"] for p in participant_dicts if p["twitch"]
            )
            runner_discord = ", ".join(
                p["discord"] for p in participant_dicts if p["discord"]
            )
            runner_twitter = ", ".join(
                p["twitter"] for p in participant_dicts if p["twitter"]
            )
        else:
            runner_display = runner_twitch = runner_discord = runner_twitter = ""

        runner_submission_ids = [
            p["submission_id"]
            for p in participant_dicts
            if p["submission_id"] is not None
        ]

        # --- Union incentive texts from all race-linked submissions ---
        incentives = _union_incentives(participant_subs)

        # --- Other submission-level fields from primary submission only ---
        primary_sub: OengusSubmission | None = participant_subs[0] if participant_subs else None
        commentator = _find_answer_by_keyword(primary_sub, ["comentarista", "commentator", "commentary"]) if primary_sub else ""
        upload_speed = _find_answer_by_keyword(primary_sub, ["upload", "velocidad", "subida", "speed"]) if primary_sub else ""
        pronouns = _find_answer_by_keyword(primary_sub, ["pronoun", "pronombres"]) if primary_sub else ""
        show_cam = _find_answer_by_keyword(primary_sub, ["cam", "avatar", "cámara"]) if primary_sub else ""
        runner_comments = _find_answer_by_keyword(primary_sub, ["comentario", "comment", "comentaris"]) if primary_sub else ""

        rows.append(CrossReferenceRow(
            scheduled=item.scheduled,
            game=item.game,
            category=item.category,
            estimate=item.estimate_str,
            platform=item.platform,
            players=item.players,
            runner_display=runner_display,
            runner_twitch=runner_twitch,
            runner_discord=runner_discord,
            runner_twitter=runner_twitter,
            note=item.note,
            layout=item.layout,
            stream=item.stream,
            submission_id=item.submission_id,
            category_id=item.category_id,
            incentives=incentives,
            commentator=commentator,
            upload_speed=upload_speed,
            pronouns=pronouns,
            show_cam=show_cam,
            runner_comments=runner_comments,
            participants=participant_dicts,
            runner_submission_ids=runner_submission_ids,
        ))

    return rows


def _load_incentive_annotations(output_path: str) -> dict[str, dict]:
    """Read existing Incentives Detail sheet, return {uuid: {text, category, valid_for_game, estimate, status}}."""
    if not os.path.exists(output_path):
        return {}
    try:
        wb = load_workbook(output_path, read_only=True, data_only=True)
    except Exception:
        return {}

    if "Incentives Detail" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["Incentives Detail"]
    annotations = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 15:
            continue
        uuid_val = str(row[14]) if row[14] else ""
        if not uuid_val:
            continue
        annotations[uuid_val] = {
            "text": str(row[7]) if len(row) > 7 and row[7] else "",
            "category": str(row[8]) if len(row) > 8 and row[8] else "",
            "valid_for_game": str(row[9]) if len(row) > 9 and row[9] else "",
            "estimate": str(row[10]) if len(row) > 10 and row[10] else "",
            "status": str(row[12]) if len(row) > 12 and row[12] else "",
        }
    wb.close()
    return annotations


def _load_fundraising_annotations(output_path: str) -> tuple[dict[str, dict[str, str]], set[str]]:
    """Read existing Fundraising View annotations. Handles old, new, and v3 column layouts."""
    if not os.path.exists(output_path):
        return {}, set()
    try:
        wb = load_workbook(output_path, read_only=True, data_only=True)
    except Exception:
        return {}, set()

    if "Fundraising View" not in wb.sheetnames:
        wb.close()
        return {}, set()

    ws = wb["Fundraising View"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    # v3 format: col 12 = "All Submission IDs" (inserted), annotations shift by 1
    is_v3_format = bool(header and len(header) > 11 and header[11] == "All Submission IDs")
    # v2 (legacy new) format: col 11 = "Submission ID", no "All Submission IDs"
    is_new_format = bool(header and len(header) > 20 and header[10] == "Submission ID" and not is_v3_format)

    annotations = {}
    all_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        game = str(row[1]) if len(row) > 1 and row[1] else ""
        category = str(row[2]) if len(row) > 2 and row[2] else ""
        key = f"{game}|{category}"

        if is_v3_format:
            # Priority at col 23 (index 22), shifted +1 from v2
            priority = str(row[22]) if len(row) > 22 and row[22] else ""
            contact_status = str(row[23]) if len(row) > 23 and row[23] else ""
            assigned_to = str(row[24]) if len(row) > 24 and row[24] else ""
            notes = str(row[25]) if len(row) > 25 and row[25] else ""
        elif is_new_format:
            priority = str(row[21]) if len(row) > 21 and row[21] else ""
            contact_status = str(row[22]) if len(row) > 22 and row[22] else ""
            assigned_to = str(row[23]) if len(row) > 23 and row[23] else ""
            notes = str(row[24]) if len(row) > 24 and row[24] else ""
        else:
            priority = str(row[10]) if len(row) > 10 and row[10] else ""
            contact_status = str(row[11]) if len(row) > 11 and row[11] else ""
            assigned_to = str(row[12]) if len(row) > 12 and row[12] else ""
            notes = str(row[13]) if len(row) > 13 and row[13] else ""

        if priority or contact_status or assigned_to or notes:
            annotations[key] = {
                "priority": priority,
                "contact_status": contact_status,
                "assigned_to": assigned_to,
                "notes": notes,
            }
            all_keys.add(key)

    wb.close()
    return annotations, all_keys


def _add_dropdown(ws, col_letter: str, options: str, rows: int):
    dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
    dv.error = "Please select a value from the dropdown list."
    dv.errorTitle = "Invalid Value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{rows + 1}")


def generate_spreadsheet(
    schedule_items: list[ScheduleItem],
    marathon: OengusMarathon,
    output_path: str,
):
    wb = Workbook()

    # --- Sheet 1: Schedule ---
    ws_sched = wb.active
    ws_sched.title = "Schedule"
    sched_headers = [
        "Scheduled (UTC)", "Game", "Category", "Estimate", "Platform",
        "Players", "Note", "Layout", "Stream", "Submission ID", "Category ID",
    ]
    _style_header(ws_sched, sched_headers)
    for i, item in enumerate(schedule_items):
        row = i + 2
        ws_sched.cell(row=row, column=1, value=item.scheduled.isoformat())
        ws_sched.cell(row=row, column=2, value=item.game)
        ws_sched.cell(row=row, column=3, value=item.category)
        ws_sched.cell(row=row, column=4, value=item.estimate_str)
        ws_sched.cell(row=row, column=5, value=item.platform)
        ws_sched.cell(row=row, column=6, value=item.players)
        ws_sched.cell(row=row, column=7, value=item.note)
        ws_sched.cell(row=row, column=8, value=item.layout)
        ws_sched.cell(row=row, column=9, value=item.stream)
        ws_sched.cell(row=row, column=10, value=item.submission_id)
        ws_sched.cell(row=row, column=11, value=item.category_id)
        _style_data_row(ws_sched, row, len(sched_headers), alt=(i % 2 == 1))
    _auto_width(ws_sched)

    # --- Sheet 2: Submissions ---
    ws_subs = wb.create_sheet("Submissions")
    subs_headers = [
        "Submission ID", "Runner", "Twitch", "Discord", "Twitter",
        "Game", "Console", "Category", "Estimate", "Emulated",
        "Incentives", "Commentator", "Upload Speed", "Pronouns",
        "Show Cam", "Comments",
    ]
    _style_header(ws_subs, subs_headers)
    row = 2
    for sub in marathon.submissions:
        for game in sub.games:
            for cat in game.categories:
                incentives = _find_answer_by_keyword(sub, ["incentive", "incentivos"])
                commentator = _find_answer_by_keyword(sub, ["comentarista", "commentator", "commentary"])
                upload_speed = _find_answer_by_keyword(sub, ["upload", "velocidad", "subida", "speed"])
                pronouns = _find_answer_by_keyword(sub, ["pronoun", "pronombres"])
                show_cam = _find_answer_by_keyword(sub, ["cam", "avatar", "cámara"])
                comments = _find_answer_by_keyword(sub, ["comentario", "comment", "comentaris"])

                ws_subs.cell(row=row, column=1, value=sub.id)
                ws_subs.cell(row=row, column=2, value=sub.user.display_name)
                ws_subs.cell(row=row, column=3, value=sub.user.twitch)
                ws_subs.cell(row=row, column=4, value=sub.user.discord)
                ws_subs.cell(row=row, column=5, value=sub.user.twitter)
                ws_subs.cell(row=row, column=6, value=game.name)
                ws_subs.cell(row=row, column=7, value=game.console)
                ws_subs.cell(row=row, column=8, value=cat.name)
                ws_subs.cell(row=row, column=9, value=cat.estimate)
                ws_subs.cell(row=row, column=10, value="Yes" if game.emulated else "No")
                ws_subs.cell(row=row, column=11, value=incentives)
                ws_subs.cell(row=row, column=12, value=commentator)
                ws_subs.cell(row=row, column=13, value=upload_speed)
                ws_subs.cell(row=row, column=14, value=pronouns)
                ws_subs.cell(row=row, column=15, value=show_cam)
                ws_subs.cell(row=row, column=16, value=comments)
                _style_data_row(ws_subs, row, len(subs_headers), alt=(row % 2 == 0))
                row += 1
    _auto_width(ws_subs)

    # --- Sheet 3: Cross-Reference ---
    ws_xref = wb.create_sheet("Cross-Reference")
    xref_headers = [
        "Scheduled (UTC)", "Game", "Category", "Estimate", "Platform",
        "Players", "Runner", "Twitch", "Discord", "Twitter",
        "Note", "Layout", "Stream", "Submission ID", "Category ID",
        "Incentives", "Commentator", "Upload Speed", "Pronouns",
        "Show Cam", "Runner Comments", "Participants JSON",
    ]
    _style_header(ws_xref, xref_headers)
    xref_rows = _build_cross_reference(schedule_items, marathon.submissions)
    for i, xr in enumerate(xref_rows):
        row = i + 2
        missing = xr.participants and not xr.incentives
        ws_xref.cell(row=row, column=1, value=xr.scheduled.isoformat())
        ws_xref.cell(row=row, column=2, value=xr.game)
        ws_xref.cell(row=row, column=3, value=xr.category)
        ws_xref.cell(row=row, column=4, value=xr.estimate)
        ws_xref.cell(row=row, column=5, value=xr.platform)
        ws_xref.cell(row=row, column=6, value=xr.players)
        ws_xref.cell(row=row, column=7, value=xr.runner_display)
        ws_xref.cell(row=row, column=8, value=xr.runner_twitch)
        ws_xref.cell(row=row, column=9, value=xr.runner_discord)
        ws_xref.cell(row=row, column=10, value=xr.runner_twitter)
        ws_xref.cell(row=row, column=11, value=xr.note)
        ws_xref.cell(row=row, column=12, value=xr.layout)
        ws_xref.cell(row=row, column=13, value=xr.stream)
        ws_xref.cell(row=row, column=14, value=xr.submission_id)
        ws_xref.cell(row=row, column=15, value=xr.category_id)
        ws_xref.cell(row=row, column=16, value=xr.incentives)
        ws_xref.cell(row=row, column=17, value=xr.commentator)
        ws_xref.cell(row=row, column=18, value=xr.upload_speed)
        ws_xref.cell(row=row, column=19, value=xr.pronouns)
        ws_xref.cell(row=row, column=20, value=xr.show_cam)
        ws_xref.cell(row=row, column=21, value=xr.runner_comments)
        ws_xref.cell(row=row, column=22, value=json.dumps(xr.participants, ensure_ascii=False))
        _style_data_row(ws_xref, row, len(xref_headers), alt=(i % 2 == 1), missing=missing)
    ws_xref.column_dimensions["V"].hidden = True
    _auto_width(ws_xref)

    # --- Sheet 4: Incentives Detail ---
    ws_inc = wb.create_sheet("Incentives Detail")
    inc_headers = [
        "Scheduled (UTC)", "Game", "Category", "Stream",
        "Runner", "Twitch", "Discord",
        "Incentive Text", "Incentive Category", "Valid for Game",
        "Incentive Estimate", "Needs Approval", "Status",
        "Submission ID", "UUID", "Participants JSON",
    ]
    _style_header(ws_inc, inc_headers)

    existing_inc = _load_incentive_annotations(output_path)

    runner_games_map: dict[str, list[str]] = {}
    for sub in marathon.submissions:
        runner_games_map[str(sub.id)] = [g.name for g in sub.games]

    inc_rows = build_incentive_rows(xref_rows, existing_inc, runner_games_map)

    for i, ir in enumerate(inc_rows):
        row = i + 2
        ws_inc.cell(row=row, column=1, value=ir.scheduled.isoformat())
        ws_inc.cell(row=row, column=2, value=ir.game)
        ws_inc.cell(row=row, column=3, value=ir.category)
        ws_inc.cell(row=row, column=4, value=ir.stream)
        ws_inc.cell(row=row, column=5, value=ir.runner_display)
        ws_inc.cell(row=row, column=6, value=ir.runner_twitch)
        ws_inc.cell(row=row, column=7, value=ir.runner_discord)
        ws_inc.cell(row=row, column=8, value=ir.incentive_text)
        ws_inc.cell(row=row, column=9, value=ir.incentive_category)
        ws_inc.cell(row=row, column=10, value=ir.valid_for_game)
        ws_inc.cell(row=row, column=11, value=ir.incentive_estimate)
        ws_inc.cell(row=row, column=12, value=ir.needs_approval)
        ws_inc.cell(row=row, column=13, value=ir.status)
        ws_inc.cell(row=row, column=14, value=ir.submission_id)
        ws_inc.cell(row=row, column=15, value=ir.row_uuid)
        ws_inc.cell(row=row, column=16, value=json.dumps(ir.participants, ensure_ascii=False))
        _style_data_row(ws_inc, row, len(inc_headers), alt=(i % 2 == 1))

    if inc_rows:
        _add_dropdown(ws_inc, "I", "Reward,Poll-Bid War,Target", len(inc_rows))
        _add_dropdown(ws_inc, "J", "Yes,No,Needs Review", len(inc_rows))
        _add_dropdown(ws_inc, "M", "To-Do,In Review,Needs Information,Approved,Removed", len(inc_rows))

    ws_inc.column_dimensions["O"].hidden = True
    ws_inc.column_dimensions["P"].hidden = True
    _auto_width(ws_inc)

    # --- Sheet 5: Fundraising View ---
    ws_fund = wb.create_sheet("Fundraising View")
    fund_headers = [
        "Scheduled (UTC)", "Game", "Category", "Estimate", "Stream",
        "Runner", "Twitch", "Discord", "Incentives", "Commentator",
        "Submission ID", "All Submission IDs",
        "Incentive Count", "Approved", "To-Do", "Needs Info", "In Review",
        "Has Reward", "Has Poll/Bid War", "Has Target",
        "Any Needs Approval", "Contact Needed",
        "Priority", "Contact Status", "Assigned To", "Notes",
    ]
    _style_header(ws_fund, fund_headers)

    existing_fund, all_annotated_keys = _load_fundraising_annotations(output_path)

    fund_row = 2
    current_keys = set()
    for i, xr in enumerate(xref_rows):
        key = f"{xr.game}|{xr.category}"
        current_keys.add(key)
        ann = existing_fund.get(key, {})
        sub_id = xr.submission_id or ""
        all_sub_ids = ", ".join(xr.runner_submission_ids) if xr.runner_submission_ids else sub_id

        ws_fund.cell(row=fund_row, column=1, value=xr.scheduled.isoformat())
        ws_fund.cell(row=fund_row, column=2, value=xr.game)
        ws_fund.cell(row=fund_row, column=3, value=xr.category)
        ws_fund.cell(row=fund_row, column=4, value=xr.estimate)
        ws_fund.cell(row=fund_row, column=5, value=xr.stream)
        ws_fund.cell(row=fund_row, column=6, value=xr.runner_display)
        ws_fund.cell(row=fund_row, column=7, value=xr.runner_twitch)
        ws_fund.cell(row=fund_row, column=8, value=xr.runner_discord)
        ws_fund.cell(row=fund_row, column=9, value=xr.incentives)
        ws_fund.cell(row=fund_row, column=10, value=xr.commentator)
        ws_fund.cell(row=fund_row, column=11, value=sub_id)
        ws_fund.cell(row=fund_row, column=12, value=all_sub_ids)

        detail_sheet = "'Incentives Detail'"
        # COUNTIF formulas still reference col K (Submission ID, primary)
        k_ref = f"K{fund_row}"

        ws_fund.cell(row=fund_row, column=13).value = f'=COUNTIF({detail_sheet}!N:N, {k_ref})'
        ws_fund.cell(row=fund_row, column=14).value = f'=COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!M:M, "Approved")'
        ws_fund.cell(row=fund_row, column=15).value = f'=COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!M:M, "To-Do")'
        ws_fund.cell(row=fund_row, column=16).value = f'=COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!M:M, "Needs Information")'
        ws_fund.cell(row=fund_row, column=17).value = f'=COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!M:M, "In Review")'
        ws_fund.cell(row=fund_row, column=18).value = f'=IF(COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!I:I, "Reward")>0, "Yes", "")'
        ws_fund.cell(row=fund_row, column=19).value = f'=IF(COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!I:I, "Poll-Bid War")>0, "Yes", "")'
        ws_fund.cell(row=fund_row, column=20).value = f'=IF(COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!I:I, "Target")>0, "Yes", "")'
        ws_fund.cell(row=fund_row, column=21).value = f'=IF(COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!L:L, "Yes")>0, "Yes", "No")'
        ws_fund.cell(row=fund_row, column=22).value = (
            f'=IF(OR('
            f'COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!M:M, "Needs Information")>0,'
            f'COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!M:M, "To-Do")>0,'
            f'COUNTIFS({detail_sheet}!N:N, {k_ref}, {detail_sheet}!L:L, "Yes")>0'
            f'), "Yes", "No")'
        )

        ws_fund.cell(row=fund_row, column=23, value=ann.get("priority", ""))
        ws_fund.cell(row=fund_row, column=24, value=ann.get("contact_status", ""))
        ws_fund.cell(row=fund_row, column=25, value=ann.get("assigned_to", ""))
        ws_fund.cell(row=fund_row, column=26, value=ann.get("notes", ""))

        _style_data_row(ws_fund, fund_row, len(fund_headers), alt=(i % 2 == 1))
        fund_row += 1

    data_end = fund_row - 1
    if data_end >= 2:
        ws_fund.conditional_formatting.add(
            f"A2:Z{data_end}",
            FormulaRule(
                formula=[f'AND(LEN(TRIM($I2))=0, $F2<>"")'],
                fill=MISSING_FILL,
            )
        )

    ws_fund.column_dimensions["K"].hidden = True
    ws_fund.column_dimensions["L"].hidden = True
    _auto_width(ws_fund)

    # --- Sheet 6: Marathon Info ---
    ws_info = wb.create_sheet("Marathon Info")
    info_data = [
        ("Marathon Name", marathon.name),
        ("Marathon ID", marathon.id),
        ("Start Date", marathon.start_date.isoformat()),
        ("End Date", marathon.end_date.isoformat()),
        ("Submissions End", marathon.submissions_end_date.isoformat() if marathon.submissions_end_date else "N/A"),
        ("Total Submissions", len(marathon.submissions)),
        ("Total Schedule Items", len(schedule_items)),
        ("Matched Runs", sum(1 for xr in xref_rows if xr.participants)),
        ("Total Incentives", len(inc_rows)),
        ("Runs with Incentives", len(set(ir.submission_id for ir in inc_rows if ir.submission_id))),
    ]
    for i, (label, value) in enumerate(info_data):
        ws_info.cell(row=i + 1, column=1, value=label).font = Font(bold=True)
        ws_info.cell(row=i + 1, column=2, value=str(value))
    ws_info.column_dimensions["A"].width = 20
    ws_info.column_dimensions["B"].width = 40

    wb.save(output_path)
    orphaned = all_annotated_keys - current_keys
    return wb, orphaned
